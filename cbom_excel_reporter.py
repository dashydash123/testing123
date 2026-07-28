#!/usr/bin/env python3
"""
CBOM -> Excel Reporter
======================

Reads two CycloneDX CBOM files (SCANOSS CBOM + Internal Tool CBOM) and produces
two independent Excel reports, each with three worksheets:

    1. Crypto Assets      - one row per component
    2. Occurrences        - one row per evidence.occurrences entry
    3. Identity Evidence  - one row per evidence.identity[].methods[] entry

Outputs:
    SCANOSS_CBOM_Output.xlsx
    Internal_CBOM_Output.xlsx

Requires: openpyxl   (pip install openpyxl)
GUI:      Tkinter (stdlib) - CustomTkinter used automatically if installed.

CLI mode (no GUI):
    python cbom_excel_reporter.py --cli <scanoss.json> <internal.json> <outdir>
"""

import json
import os
import re
import sys
import threading

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_TITLE = "CBOM Excel Reporter"

# --------------------------------------------------------------------------
# Sheet definitions
# --------------------------------------------------------------------------

CRYPTO_ASSET_COLUMNS = [
    "Asset Name",
    "bom-ref",
    "type",
    "name",
    "description",
    "assetType",
    "primitive",
    "parameterSetIdentifier",
    "executionEnvironment",
    "implementationPlatform",
    "cryptoFunctions",
    "oid",
    "relatedCryptoMaterialType",
    "scanossCryptoFunction",
    "Occurrence Count",
    "Identity Evidence Count",
]

OCCURRENCE_COLUMNS = [
    "Asset Name",
    "bom-ref",
    "Location",
    "Line",
    "Additional Context",
]

IDENTITY_COLUMNS = [
    "Asset Name",
    "bom-ref",
    "Field",
    "Identity Confidence",
    "Technique",
    "Method Confidence",
    "Method Value",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

MAX_COL_WIDTH = 70
MIN_COL_WIDTH = 12

# --------------------------------------------------------------------------
# Tolerant JSON loading
# --------------------------------------------------------------------------

# Matches an unquoted scalar that looks like a bom-ref (contains / or @ or :)
_BARE_TOKEN = r"[A-Za-z0-9_.\-]+(?:[/@:][A-Za-z0-9_.\-@/:+]*)+"

_BARE_AFTER_COLON = re.compile(r'(:\s*)(' + _BARE_TOKEN + r')(\s*,?\s*)$')
_BARE_ARRAY_ITEM = re.compile(r'^(\s*)(' + _BARE_TOKEN + r')(\s*,?\s*)$')


def _repair_json(text):
    """
    Repair the most common defect seen in hand-edited CBOM samples: bom-ref /
    dependency-ref values emitted without surrounding quotes, e.g.

        "bom-ref": crypto/algorithm/eddsa@1.3.101.112,

    Returns (repaired_text, repair_count).
    """
    repaired_lines = []
    count = 0
    for line in text.splitlines():
        new_line, n = _BARE_AFTER_COLON.subn(r'\1"\2"\3', line)
        if n == 0:
            new_line, n = _BARE_ARRAY_ITEM.subn(r'\1"\2"\3', line)
        count += n
        repaired_lines.append(new_line)
    return "\n".join(repaired_lines), count


def load_cbom(path):
    """Load a CBOM file, repairing malformed unquoted refs if necessary.

    Returns (data_dict, notes_list).
    """
    notes = []
    with open(path, "r", encoding="utf-8-sig") as fh:
        raw = fh.read()

    try:
        return json.loads(raw), notes
    except json.JSONDecodeError as first_error:
        repaired, n = _repair_json(raw)
        if n:
            try:
                data = json.loads(repaired)
                notes.append(
                    "%s: repaired %d unquoted JSON value(s) before parsing."
                    % (os.path.basename(path), n)
                )
                return data, notes
            except json.JSONDecodeError as second_error:
                raise ValueError(
                    "%s is not valid JSON (line %d, col %d): %s"
                    % (
                        os.path.basename(path),
                        second_error.lineno,
                        second_error.colno,
                        second_error.msg,
                    )
                )
        raise ValueError(
            "%s is not valid JSON (line %d, col %d): %s"
            % (
                os.path.basename(path),
                first_error.lineno,
                first_error.colno,
                first_error.msg,
            )
        )


# --------------------------------------------------------------------------
# Extraction
# --------------------------------------------------------------------------

def _flatten(value):
    """Render lists/dicts as a readable single cell value."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, tuple)):
        return "; ".join(_flatten(v) for v in value if v is not None)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return value


def _scanoss_crypto_function(component):
    """Pull scanoss:cryptoFunction (and any other scanoss crypto property)."""
    values = []
    for prop in component.get("properties") or []:
        if not isinstance(prop, dict):
            continue
        name = (prop.get("name") or "").strip()
        if name.lower() in ("scanoss:cryptofunction", "scanoss:crypto-function"):
            values.append(str(prop.get("value", "")))
    return "; ".join(v for v in values if v)


def extract(data):
    """Return (assets_rows, occurrence_rows, identity_rows) from a CBOM dict."""
    assets, occurrences, identities = [], [], []

    components = data.get("components") or []
    if not isinstance(components, list):
        components = []

    for component in components:
        if not isinstance(component, dict):
            continue

        name = component.get("name", "") or ""
        bom_ref = component.get("bom-ref", "") or ""

        crypto = component.get("cryptoProperties") or {}
        algo = crypto.get("algorithmProperties") or {}
        material = crypto.get("relatedCryptoMaterialProperties") or {}

        evidence = component.get("evidence") or {}
        occ_list = evidence.get("occurrences") or []
        id_list = evidence.get("identity") or []
        # CycloneDX allows evidence.identity to be a single object
        if isinstance(id_list, dict):
            id_list = [id_list]

        # ---- Sheet 2: Occurrences -------------------------------------
        occ_count = 0
        for occ in occ_list:
            if not isinstance(occ, dict):
                continue
            occ_count += 1
            occurrences.append([
                name,
                bom_ref,
                occ.get("location", "") or "",
                occ.get("line", "") if occ.get("line") is not None else "",
                occ.get("additionalContext", "") or "",
            ])

        # ---- Sheet 3: Identity Evidence -------------------------------
        id_count = 0
        for ident in id_list:
            if not isinstance(ident, dict):
                continue
            field = ident.get("field", "") or ""
            confidence = ident.get("confidence", "")
            methods = ident.get("methods") or []
            if isinstance(methods, dict):
                methods = [methods]
            if not methods:
                id_count += 1
                identities.append([name, bom_ref, field, confidence, "", "", ""])
                continue
            for method in methods:
                if not isinstance(method, dict):
                    continue
                id_count += 1
                identities.append([
                    name,
                    bom_ref,
                    field,
                    confidence,
                    method.get("technique", "") or "",
                    method.get("confidence", ""),
                    method.get("value", "") or "",
                ])

        # ---- Sheet 1: Crypto Assets -----------------------------------
        assets.append([
            name,                                            # Asset Name
            bom_ref,                                         # bom-ref
            component.get("type", "") or "",                 # type
            name,                                            # name
            component.get("description", "") or "",          # description
            crypto.get("assetType", "") or "",               # assetType
            algo.get("primitive", "") or "",                 # primitive
            algo.get("parameterSetIdentifier", "") or "",    # parameterSetIdentifier
            algo.get("executionEnvironment", "") or "",      # executionEnvironment
            algo.get("implementationPlatform", "") or "",    # implementationPlatform
            _flatten(algo.get("cryptoFunctions")),           # cryptoFunctions
            crypto.get("oid", "") or "",                     # oid
            material.get("type", "") or "",                  # relatedCryptoMaterialType
            _scanoss_crypto_function(component),             # scanossCryptoFunction
            occ_count,                                       # Occurrence Count
            id_count,                                        # Identity Evidence Count
        ])

    return assets, occurrences, identities


# --------------------------------------------------------------------------
# Excel writing
# --------------------------------------------------------------------------

def _write_sheet(ws, columns, rows):
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (
        get_column_letter(len(columns)),
        max(len(rows) + 1, 1),
    )

    for idx, column in enumerate(columns, start=1):
        longest = len(str(column))
        for row in rows:
            value = row[idx - 1]
            if value is None:
                continue
            longest = max(longest, len(str(value)))
        width = min(max(longest + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 28


def write_report(path, assets, occurrences, identities):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Crypto Assets"
    _write_sheet(ws1, CRYPTO_ASSET_COLUMNS, assets)

    ws2 = wb.create_sheet("Occurrences")
    _write_sheet(ws2, OCCURRENCE_COLUMNS, occurrences)

    ws3 = wb.create_sheet("Identity Evidence")
    _write_sheet(ws3, IDENTITY_COLUMNS, identities)

    wb.save(path)


def build_report(cbom_path, output_path):
    """Parse one CBOM and write one Excel report. Returns a summary dict."""
    data, notes = load_cbom(cbom_path)
    assets, occurrences, identities = extract(data)
    write_report(output_path, assets, occurrences, identities)
    return {
        "output": output_path,
        "assets": len(assets),
        "occurrences": len(occurrences),
        "identities": len(identities),
        "notes": notes,
    }


def process(scanoss_path, internal_path, output_dir, progress=None):
    """Run both reports. `progress` is an optional callable(pct, message)."""
    def tick(pct, msg):
        if progress:
            progress(pct, msg)

    os.makedirs(output_dir, exist_ok=True)
    results = []

    tick(10, "Reading SCANOSS CBOM...")
    results.append(
        build_report(scanoss_path, os.path.join(output_dir, "SCANOSS_CBOM_Output.xlsx"))
    )

    tick(55, "Reading Internal Tool CBOM...")
    results.append(
        build_report(internal_path, os.path.join(output_dir, "Internal_CBOM_Output.xlsx"))
    )

    tick(100, "Done.")
    return results


# --------------------------------------------------------------------------
# GUI
# --------------------------------------------------------------------------

def launch_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title(APP_TITLE)
    root.geometry("720x360")
    root.minsize(660, 340)
    root.configure(bg="#F4F5F7")

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass
    style.configure("TFrame", background="#F4F5F7")
    style.configure("TLabel", background="#F4F5F7", font=("Segoe UI", 10))
    style.configure("Header.TLabel", font=("Segoe UI Semibold", 15))
    style.configure("Sub.TLabel", foreground="#5A6472", font=("Segoe UI", 9))
    style.configure("TButton", font=("Segoe UI", 9), padding=5)
    style.configure("Go.TButton", font=("Segoe UI Semibold", 10), padding=8)

    scanoss_var = tk.StringVar()
    internal_var = tk.StringVar()
    output_var = tk.StringVar(value=os.getcwd())
    status_var = tk.StringVar(value="Select both CBOM files to begin.")

    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)
    frame.columnconfigure(1, weight=1)

    ttk.Label(frame, text="CBOM Excel Reporter", style="Header.TLabel").grid(
        row=0, column=0, columnspan=3, sticky="w"
    )
    ttk.Label(
        frame,
        text="Generates SCANOSS_CBOM_Output.xlsx and Internal_CBOM_Output.xlsx "
             "(Crypto Assets / Occurrences / Identity Evidence).",
        style="Sub.TLabel",
        wraplength=650,
    ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(2, 16))

    def pick_file(var):
        path = filedialog.askopenfilename(
            title="Select CBOM file",
            filetypes=[("CBOM / JSON", "*.json *.txt *.cdx.json"), ("All files", "*.*")],
        )
        if path:
            var.set(path)

    def pick_dir():
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            output_var.set(path)

    def add_row(row, label, var, command):
        ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", pady=6)
        entry = ttk.Entry(frame, textvariable=var)
        entry.grid(row=row, column=1, sticky="ew", padx=10, pady=6)
        ttk.Button(frame, text="Browse...", command=command).grid(
            row=row, column=2, sticky="e", pady=6
        )

    add_row(2, "SCANOSS CBOM", scanoss_var, lambda: pick_file(scanoss_var))
    add_row(3, "Internal Tool CBOM", internal_var, lambda: pick_file(internal_var))
    add_row(4, "Output folder", output_var, pick_dir)

    progress = ttk.Progressbar(frame, mode="determinate", maximum=100)
    progress.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(20, 6))

    status = ttk.Label(frame, textvariable=status_var, style="Sub.TLabel", wraplength=650)
    status.grid(row=6, column=0, columnspan=3, sticky="w")

    run_button = ttk.Button(frame, text="Process", style="Go.TButton")
    run_button.grid(row=7, column=2, sticky="e", pady=(14, 0))

    def set_progress(pct, message):
        root.after(0, lambda: (progress.configure(value=pct), status_var.set(message)))

    def worker(scanoss, internal, outdir):
        try:
            results = process(scanoss, internal, outdir, progress=set_progress)
        except Exception as exc:                                # noqa: BLE001
            root.after(0, lambda: on_error(exc))
        else:
            root.after(0, lambda: on_success(results, outdir))

    def on_error(exc):
        progress.configure(value=0)
        status_var.set("Failed: %s" % exc)
        run_button.configure(state="normal")
        messagebox.showerror(APP_TITLE, str(exc))

    def on_success(results, outdir):
        run_button.configure(state="normal")
        lines = []
        for res in results:
            lines.append(
                "%s\n    %d assets, %d occurrences, %d identity evidence rows"
                % (
                    os.path.basename(res["output"]),
                    res["assets"],
                    res["occurrences"],
                    res["identities"],
                )
            )
            lines.extend("    Note: " + n for n in res["notes"])
        status_var.set("Reports written to %s" % outdir)
        messagebox.showinfo(APP_TITLE, "Done.\n\n" + "\n".join(lines))

    def on_run():
        scanoss, internal, outdir = (
            scanoss_var.get().strip(),
            internal_var.get().strip(),
            output_var.get().strip(),
        )
        if not scanoss or not os.path.isfile(scanoss):
            messagebox.showwarning(APP_TITLE, "Select a valid SCANOSS CBOM file.")
            return
        if not internal or not os.path.isfile(internal):
            messagebox.showwarning(APP_TITLE, "Select a valid Internal Tool CBOM file.")
            return
        if not outdir:
            messagebox.showwarning(APP_TITLE, "Select an output folder.")
            return

        run_button.configure(state="disabled")
        progress.configure(value=0)
        status_var.set("Processing...")
        threading.Thread(
            target=worker, args=(scanoss, internal, outdir), daemon=True
        ).start()

    run_button.configure(command=on_run)
    root.mainloop()


# --------------------------------------------------------------------------

def main():
    if len(sys.argv) > 1 and sys.argv[1] == "--cli":
        if len(sys.argv) != 5:
            print("usage: %s --cli <scanoss.json> <internal.json> <outdir>" % sys.argv[0])
            return 2
        for res in process(sys.argv[2], sys.argv[3], sys.argv[4]):
            print(
                "%s -> %d assets, %d occurrences, %d identity rows"
                % (res["output"], res["assets"], res["occurrences"], res["identities"])
            )
            for note in res["notes"]:
                print("   note: %s" % note)
        return 0
    launch_gui()
    return 0


if __name__ == "__main__":
    sys.exit(main())
